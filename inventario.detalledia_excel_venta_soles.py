import os
import hashlib
import mysql.connector
from openpyxl import load_workbook
from datetime import datetime

# Configuración de conexión a la base de datos
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',         # 👈 Cambia esto
    'password': 'Sistemas321',  # 👈 Cambia esto
    'database': 'gasolinas1',
    'charset': 'utf8mb4',
    'collation': 'utf8mb4_general_ci'
}

# Fila del producto por tipo
PRODUCTOS_FILA = {
    7: 1,  # GLP
    8: 2,  # PREMIUM
    9: 3,  # REGULAR
    10: 4  # DB5
}

def obtener_hash_archivo(nombre_archivo):
    hasher = hashlib.sha256()
    with open(nombre_archivo, 'rb') as f:
        hasher.update(f.read())
    return hasher.hexdigest()

def cargar_datos_hoja(sheet, fecha, estacion_id, cursor):
    for fila, producto_id in PRODUCTOS_FILA.items():
        compra = sheet[f"C{fila}"].value
        surtidor = sheet[f"D{fila}"].value
        ventas = sheet[f"G{fila}"].value
        invent_final = sheet[f"K{fila}"].value
        var_ini = sheet[f"L{fila}"].value
        var_fin = sheet[f"M{fila}"].value
        dif = sheet[f"Q{fila}"].value

        # Insertar o actualizar en detalle_dia
        cursor.execute("""
            SELECT id FROM detalle_dia
            WHERE fecha = %s AND productoid = %s AND estacionid = %s
        """, (fecha, producto_id, estacion_id))
        if cursor.fetchone():
            cursor.execute("""
                UPDATE detalle_dia
                SET compra_planta=%s, prueba_surtidor=%s, total_ventas=%s,
                    invent_final=%s, varillaje_inicial=%s, varillaje_final=%s, dif=%s
                WHERE fecha = %s AND productoid = %s AND estacionid = %s
            """, (compra, surtidor, ventas, invent_final, var_ini, var_fin, dif,
                  fecha, producto_id, estacion_id))
        else:
            cursor.execute("""
                INSERT INTO detalle_dia (
                    fecha, estacionid, productoid, compra_planta, prueba_surtidor,
                    total_ventas, invent_final, varillaje_inicial, varillaje_final, dif
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (fecha, estacion_id, producto_id, compra, surtidor, ventas,
                  invent_final, var_ini, var_fin, dif))

        # Insertar o actualizar inventario
        cantidad = ventas or 0
        precio_venta = 18.00
        venta_soles = round(cantidad * precio_venta, 3)

        cursor.execute("""
            SELECT id FROM inventario
            WHERE fecha = %s AND productoid = %s AND estacionid = %s
        """, (fecha, producto_id, estacion_id))
        if cursor.fetchone():
            cursor.execute("""
                UPDATE inventario
                SET cantidad=%s, precio_venta=%s, venta_soles=%s
                WHERE fecha = %s AND productoid = %s AND estacionid = %s
            """, (cantidad, precio_venta, venta_soles, fecha, producto_id, estacion_id))
        else:
            cursor.execute("""
                INSERT INTO inventario (
                    fecha, productoid, estacionid, cantidad, precio_venta, venta_soles
                ) VALUES (%s, %s, %s, %s, %s, %s)
            """, (fecha, producto_id, estacion_id, cantidad, precio_venta, venta_soles))

def procesar_excel(nombre_archivo, forzar=False):
    nombre_base = os.path.basename(nombre_archivo).lower()
    if "america" in nombre_base:
        estacion_id = 3
    elif "rinconada" in nombre_base:
        estacion_id = 1
    elif "porvenir" in nombre_base:
        estacion_id = 4
    else:
        print(f"⚠️ No se reconoce la estación en el archivo '{nombre_archivo}'")
        return

    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor()

    hash_archivo = obtener_hash_archivo(nombre_archivo)
    cursor.execute("SELECT id FROM archivos_procesados WHERE hash = %s", (hash_archivo,))
    ya_procesado = cursor.fetchone()

    if ya_procesado and not forzar:
        print(f"⚠️ El archivo '{nombre_archivo}' ya fue procesado. Se omite.")
        conn.close()
        return

    wb = load_workbook(filename=nombre_archivo, data_only=True)
    for hoja in wb.sheetnames:
        try:
            fecha = datetime.strptime(hoja.strip(), "%d-%m-%Y").date()
            sheet = wb[hoja]
            cargar_datos_hoja(sheet, fecha, estacion_id, cursor)
        except Exception as e:
            print(f"❌ Error procesando hoja '{hoja}': {e}")

    if not ya_procesado:
        cursor.execute("INSERT INTO archivos_procesados (nombre_archivo, hash) VALUES (%s, %s)",
                       (nombre_archivo, hash_archivo))

    conn.commit()
    cursor.close()
    conn.close()
    print(f"✅ Archivo procesado y actualizado: {nombre_archivo}")

if __name__ == "__main__":
    archivo = input("Ingrese el nombre del archivo Excel: ")
    forzar = input("¿Deseas forzar la recarga y actualizar datos existentes? (s/n): ").lower() == 's'
    if os.path.exists(archivo):
        procesar_excel(archivo, forzar=forzar)
    else:
        print("❌ Archivo no encontrado.")

